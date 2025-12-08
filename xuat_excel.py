import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
import database

def export_to_excel(month=None):
    """
    Xuất dữ liệu chi tiêu + thu nhập ra Excel.
    Nếu không truyền tháng, tự động lấy tháng hiện tại.
    """
    if not month or not isinstance(month, str):
        month = datetime.now().strftime("%m-%Y")
    month = month.strip()

    income = database.get_income_for_month(month)

    # Lấy dữ liệu chi tiêu theo tháng
    conn = database.get_conn()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(c.name, 'Khác') AS category, e.description, e.amount, e.date
            FROM expenses e
            LEFT JOIN categories c ON e.category_id = c.id
            WHERE substr(e.date, 4, 7) = ?
            ORDER BY substr(e.date,7,4) DESC, substr(e.date,4,2) DESC, substr(e.date,1,2) DESC
        """, (month,))
        expenses = cur.fetchall()
    finally:
        conn.close()

    if not income and not expenses:
        raise ValueError(f"Không có dữ liệu cho tháng {month}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tháng {month}"

    # Cấu hình style
    title_font = Font(bold=True, size=14, color="FFFFFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")
    fill_header = PatternFill(start_color="C89F6D", end_color="C89F6D", fill_type="solid")
    fill_title = PatternFill(start_color="7A5C3E", end_color="7A5C3E", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Tiêu đề
    ws.merge_cells("A1:E1")
    ws["A1"] = f"BÁO CÁO CHI TIÊU THÁNG {month}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = fill_title

    ws.append([""])
    ws.append(["Thu nhập (VND):", f"{income:,.0f}" if income else "0"])
    ws.append([""])
    ws.append(["STT", "Danh mục", "Mô tả", "Số tiền (VND)", "Ngày"])

    total_spent = 0
    for i, (cat, desc, amount, date) in enumerate(expenses, start=1):
        total_spent += amount
        ws.append([i, cat, desc, amount, date])

    ws.append([""])
    ws.append(["", "", "TỔNG CHI", total_spent])
    ws.append(["", "", "SỐ DƯ", (income or 0) - total_spent])

    # Thêm border + căn chỉnh
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = right if cell.column == 4 else center

    for cell in ws[5]:
        cell.font = bold
        cell.fill = fill_header

    # Căn chỉnh độ rộng cột
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    max_row = ws.max_row
    for c in ["C", "D"]:
        ws[f"{c}{max_row - 1}"].font = bold
        ws[f"{c}{max_row}"].font = bold

    # Lưu file, tránh trùng tên
    program_dir = os.path.dirname(os.path.abspath(__file__))
    base_filename = f"chi_tieu-{month.replace('/', '-')}.xlsx"
    save_path = os.path.join(program_dir, base_filename)
    counter = 1
    while os.path.exists(save_path):
        name, ext = os.path.splitext(base_filename)
        save_path = os.path.join(program_dir, f"{name}({counter}){ext}")
        counter += 1

    wb.save(save_path)
    return save_path
